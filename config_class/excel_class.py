

from openpyxl import load_workbook

class Excel:
    def __init__(self, file):
        self.file = file
    
    def open_file(self):
        workbook = load_workbook(self.file)
        return workbook 
    
    def show_sheet(self, workbook):
        return workbook.sheetnames
    
    def select_sheet(self, workbook, sheet):
        return  workbook[sheet]
    
    def get_max_col(self, sheet):
        return sheet.max_column
    
    def get_max_row(self, sheet):
        return sheet.max_row
    
    def get_headers(self, sheet, index, row, header=[]):
        for i in range(index, row+1):
            cell = sheet.cell(row=1, column=i)
            header.append(cell.value)
        return header

    def get_data(self, sheet, row, column, data=[]):
        for i in range(2, row+1):
            for j in range(1, column+1):
                cell = sheet.cell(row=i, column=j)
                data.append(cell.value)
        return data
    
    def dictionary_data(self, header, data, counter):
        return [(dict(zip(header, data))) for data in [data[x:x+counter] for x in range(0, len(data), counter)]]
    

