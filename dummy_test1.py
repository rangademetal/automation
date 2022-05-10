import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
# Open the workbook and select a worksheet
wb = load_workbook('excel-xlrd-sample.xlsx')
sheet = wb['Sheet1']
# List to hold dictionaries
cars_list = []
# Iterate through each row in worksheet and fetch values into dict
for row in islice(sheet.values, 1, sheet.max_row):
    cars = OrderedDict()
    cars['car-id'] = row[0]
    cars['make'] = row[1]
    cars['model'] = row[2]
    cars['miles'] = row[3]
    cars_list.append(cars)
# Serialize the list of dicts to JSON
j = json.dumps(cars_list)
# Write to file
with open('data.json', 'w') as f:
    f.write(j)